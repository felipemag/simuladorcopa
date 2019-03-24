import openpyxl
import csv
import pandas as pd
import time

def openRawData():
    wb = openpyxl.load_workbook('Data/CompleteDataset.xlsx')
    sheet = wb['CompleteDataset']
    return(sheet)

def generateData(selecao1, selecao2):

    order = ["Goleiro","Ataque","MeioCampo","Defesa"]
    
   # if selecoes[0] is None:
   #     selecao1 = {'Seleção':"Brasil",'Goleiro':[397],'Ataque':[299],'MeioCampo':[2,114,54,488,295],'Defesa':[89,63,140,124]}
#
   # if selecoes[1] is None:
   #     selecao2 = {'Seleção':"Belgica",'Goleiro':[12],'Ataque':[7,50,374],'MeioCampo':[11,274,183],'Defesa':[90,86,806,56]}

    ListaSelecoes = [selecao1,selecao2]

    sheet = openRawData()
    maxRows = sheet.max_row
    idJogador = []
    overall = []
    finishing = []
    nome =[]

    #Pega os dados do excel
    for row in range(2,maxRows):
        idJogador.append(sheet.cell(row, column=1).value)
        overall.append(sheet.cell(row, column=7).value)
        nome.append(sheet.cell(row, column=2).value)
        finishing.append(sheet.cell(row, column=23).value)

    NomeJogador = []
    forca = []
    novoId = []
    posicao= []
    selecao = []
    finalizacao = []

    for a in ListaSelecoes:
        for y in order:
            JogadoresPosicao = a[y]
            print("Aqui>>>>",JogadoresPosicao)
            for w in range(len(JogadoresPosicao)):
                for x in range(len(idJogador)):
                    if JogadoresPosicao[w] == idJogador[x]:
                        NomeJogador.append(nome[x])
                        forca.append(overall[x])
                        novoId.append(idJogador[x])
                        posicao.append(y)
                        selecao.append(a['Seleção'])
                        finalizacao.append(finishing[x]) 

    dic = {'NomeJogador':NomeJogador,'Overall':forca,'finalizacao':finalizacao,'idJogador':novoId,'Posição':posicao,'Seleção':selecao}

    df = pd.DataFrame(dic)

    df.to_csv(''+ str(selecao1['Seleção'])+str(selecao2['Seleção'])+'.csv',sep=';')

    print(df)

    print(df[(df['Posição'] == 'MeioCampo') & (df['Seleção'] == 'Brasil')][['Overall']])

    Lista = []
    for x in ListaSelecoes:
        selectSelecao = x
        selectSelecao = selectSelecao['Seleção']
        MC = df[(df['Posição'] == 'MeioCampo') & (df['Seleção'] == selectSelecao)][['Overall']].sum()
        DF = df[(df['Posição'] == 'Defesa') & (df['Seleção'] == selectSelecao)][['Overall']].sum() + MC/4
        ATK = df[(df['Posição'] == 'Ataque') & (df['Seleção'] == selectSelecao)][['Overall']].sum() + MC/4
        GK = df[(df['Posição'] == 'Goleiro') & (df['Seleção'] == selectSelecao)][['Overall']].sum() + DF/2

        time = {'nome':selectSelecao,'gk':int(GK),'df':int(DF),'meio':int(MC),'atk':int(ATK)}
        Lista.append(time)

    df = pd.DataFrame(Lista)
    writer = pd.ExcelWriter('ValoresSelecoes.xlsx', engine='xlsxwriter')
    df.to_excel(writer, sheet_name='Sheet1')
    writer.save()


def openGeneratedData():
    Lista = []
    wb = openpyxl.load_workbook('ValoresSelecoes.xlsx')
    sheet = wb['Sheet1']
    maxRows = sheet.max_row
    for i in range(2, maxRows+1):
        ATK = sheet.cell(row=i, column=2).value 
        DF = sheet.cell(row=i, column=3).value
        GK = sheet.cell(row=i, column=4).value
        MC = sheet.cell(row=i, column=5).value
        NomeSelecao = sheet.cell(row=i, column=6).value
        time = {'nome':NomeSelecao,'gk':int(GK),'df':int(DF),'meio':int(MC),'atk':int(ATK)}
        Lista.append(time)
    return (Lista[0],Lista[1])
